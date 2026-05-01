"""XLSX генератор — JSON через openpyxl."""
import json
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def generate_xlsx(out_path: Path, content_md: str | None = None,
                  content_json: str | None = None, title: str | None = None) -> None:
    """Создать .xlsx из JSON структуры.

    Формат JSON:
    {
      "sheets": [
        {"name": "Q1", "rows": [["Месяц", "Выручка"], ["Янв", 100], ["Фев", 150]]}
      ]
    }
    """
    if not content_json:
        raise ValueError("content_json is required for xlsx format")

    data = json.loads(content_json) if isinstance(content_json, str) else content_json
    sheets = data.get("sheets") or []
    if not sheets:
        raise ValueError("xlsx: sheets list is empty")

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_data in sheets:
        name = sheet_data.get("name", "Sheet")[:31]  # лимит Excel
        rows = sheet_data.get("rows") or []
        ws = wb.create_sheet(title=name)

        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = HEADER_ALIGN

        # Auto-width
        if rows:
            for c_idx in range(1, len(rows[0]) + 1):
                col_letter = ws.cell(row=1, column=c_idx).column_letter
                max_len = max(
                    (len(str(r[c_idx-1])) for r in rows if c_idx-1 < len(r)),
                    default=10,
                )
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(str(out_path))
    log.info("xlsx generated: %s (%d sheets)", out_path, len(sheets))
